"""Parse client training-record Excel exports and upsert employees + trainings."""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from api.models import Employee, Training
from api.utils import calc_expiration, log_audit

HEADER_ALIASES = {
    'id no.': 'employee_id',
    'id no': 'employee_id',
    'employee id': 'employee_id',
    'full name': 'full_name',
    'employment': 'employment_status',
    'employment status': 'employment_status',
    'date hired': 'hire_date',
    'team': 'team',
    'line': 'line',
    'classification': 'process_classification',
    'training title': 'title',
    'category': 'category',
    'number of takes': 'take',
    'take': 'take',
    'validity (year / months)': 'validity',
    'validity': 'validity',
    'training date': 'training_date',
    'trainer': 'trainer',
    'expiration': 'expiration_date',
    'remarks': 'remarks',
    'factory (1st / 2nd)': 'factory',
    'factory': 'factory',
}

CLASSIFICATION_MAP = {
    'SENSING': 'Sensing',
    'NON SENSING': 'Non-sensing',
    'NON-SENSING': 'Non-sensing',
    'NONSENSING': 'Non-sensing',
}

TAKE_RE = re.compile(r'(\d+)')
VALIDITY_RE = re.compile(
    r'^\s*([\d.]+)\s*(year|years|yr|yrs|month|months|mo|mos)?\s*$',
    re.IGNORECASE,
)


def _norm_header(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_excel_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (1899-12-30 epoch used by Excel/Windows)
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    text = _cell_str(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(text))).date()
    except (OverflowError, ValueError):
        return None


def parse_full_name(full_name: str):
    full_name = (full_name or '').strip()
    if not full_name:
        return '', '', '', ''

    if ',' in full_name:
        last_name, rest = full_name.split(',', 1)
        parts = rest.strip().split()
        first_name = parts[0] if parts else ''
        middle_initial = ''
        if len(parts) > 1:
            middle_initial = parts[1][0].upper()
        return last_name.strip(), first_name, middle_initial, full_name

    parts = full_name.split()
    if len(parts) == 1:
        return parts[0], parts[0], '', full_name
    first_name = parts[0]
    last_name = parts[-1]
    middle_initial = parts[1][0].upper() if len(parts) > 2 else ''
    return last_name, first_name, middle_initial, full_name


def parse_take(value) -> int:
    text = _cell_str(value).upper()
    if not text:
        return 1
    match = TAKE_RE.search(text)
    if match:
        take = int(match.group(1))
        return take if take > 0 else 1
    return 1


def parse_validity_months(value):
    text = _cell_str(value)
    if not text:
        return None
    match = VALIDITY_RE.match(text)
    if not match:
        return None
    amount = Decimal(match.group(1))
    unit = (match.group(2) or 'months').lower()
    if unit.startswith('y'):
        amount = amount * 12
    try:
        return float(amount)
    except (InvalidOperation, ValueError):
        return None


def map_classification(value: str) -> str:
    raw = _cell_str(value)
    if not raw:
        return ''
    mapped = CLASSIFICATION_MAP.get(raw.upper())
    if mapped:
        return mapped
    # Title-case common variants already matching UI options
    for option in ('Beginner', 'Basic', 'Expert', 'Advanced', 'Non-sensing', 'Sensing'):
        if raw.lower() == option.lower():
            return option
    return raw


def _map_headers(header_row) -> dict:
    mapping = {}
    for idx, cell in enumerate(header_row):
        key = HEADER_ALIASES.get(_norm_header(cell))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _row_dict(values, header_map) -> dict:
    data = {}
    for key, idx in header_map.items():
        data[key] = values[idx] if idx < len(values) else None
    return data


def read_import_rows(file_obj) -> list[dict]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError('The Excel file is empty.')

        header_map = _map_headers(header_row)
        required = {'employee_id', 'title', 'training_date'}
        missing = required - set(header_map)
        if missing:
            labels = {
                'employee_id': 'ID NO.',
                'title': 'TRAINING TITLE',
                'training_date': 'TRAINING DATE',
            }
            raise ValueError(
                'Missing required column(s): '
                + ', '.join(labels[m] for m in sorted(missing))
            )

        parsed = []
        for excel_row_num, values in enumerate(rows_iter, start=2):
            if values is None or all(v is None or str(v).strip() == '' for v in values):
                continue
            raw = _row_dict(list(values), header_map)
            employee_id = _cell_str(raw.get('employee_id'))
            title = _cell_str(raw.get('title'))
            if not employee_id and not title:
                continue
            parsed.append({'row': excel_row_num, **raw, 'employee_id': employee_id, 'title': title})
        return parsed
    finally:
        wb.close()


def import_training_records(file_obj, user) -> dict:
    rows = read_import_rows(file_obj)
    if not rows:
        raise ValueError('No data rows found in the Excel file.')

    errors = []
    employee_payloads = {}  # employee_id -> attrs (last wins)

    for item in rows:
        emp_code = item['employee_id']
        if not emp_code:
            errors.append({'row': item['row'], 'error': 'Missing ID NO.'})
            continue

        full_name = _cell_str(item.get('full_name'))
        last_name, first_name, middle_initial, normalized_name = parse_full_name(full_name)
        if not normalized_name:
            normalized_name = emp_code
            last_name = emp_code
            first_name = emp_code

        employee_payloads[emp_code] = {
            'employee_id': emp_code,
            'last_name': last_name or emp_code,
            'first_name': first_name or emp_code,
            'middle_initial': middle_initial,
            'full_name': normalized_name,
            'factory': _cell_str(item.get('factory')),
            'line': _cell_str(item.get('line')),
            'team': _cell_str(item.get('team')),
            'employment_status': _cell_str(item.get('employment_status')),
            'hire_date': parse_excel_date(item.get('hire_date')),
            'status': 'active',
        }

    existing = {
        e.employee_id: e
        for e in Employee.objects.filter(employee_id__in=employee_payloads.keys())
    }

    to_create = []
    to_update = []
    for code, attrs in employee_payloads.items():
        emp = existing.get(code)
        if emp is None:
            to_create.append(Employee(**attrs))
            continue

        changed = False
        for field, value in attrs.items():
            if field == 'employee_id':
                continue
            if value in (None, '') and field != 'hire_date':
                continue
            if getattr(emp, field) != value:
                setattr(emp, field, value)
                changed = True
        if changed:
            to_update.append(emp)

    with transaction.atomic():
        if to_create:
            Employee.objects.bulk_create(to_create, batch_size=500)
        if to_update:
            Employee.objects.bulk_update(
                to_update,
                [
                    'last_name', 'first_name', 'middle_initial', 'full_name',
                    'factory', 'line', 'team', 'employment_status', 'hire_date', 'status',
                ],
                batch_size=500,
            )

        employees = {
            e.employee_id: e
            for e in Employee.objects.filter(employee_id__in=employee_payloads.keys())
        }

        existing_keys = {
            (t.employee_id, t.title, t.training_date, t.take)
            for t in Training.objects.filter(
                is_archived=False,
                employee_id__in=[e.id for e in employees.values()],
            ).only('employee_id', 'title', 'training_date', 'take')
        }

        trainings_to_create = []
        skipped_duplicates = 0
        seen_in_file = set()

        for item in rows:
            emp_code = item['employee_id']
            title = item['title']
            if not emp_code or not title:
                if emp_code or title:
                    errors.append({
                        'row': item['row'],
                        'error': 'Missing ID NO. or TRAINING TITLE',
                    })
                continue

            employee = employees.get(emp_code)
            if not employee:
                errors.append({'row': item['row'], 'error': f'Employee {emp_code} could not be created'})
                continue

            training_date = parse_excel_date(item.get('training_date'))
            if not training_date:
                errors.append({'row': item['row'], 'error': 'Invalid or missing TRAINING DATE'})
                continue

            take = parse_take(item.get('take'))
            key = (employee.id, title, training_date, take)
            if key in existing_keys or key in seen_in_file:
                skipped_duplicates += 1
                continue
            seen_in_file.add(key)

            validity_months = parse_validity_months(item.get('validity'))
            expiration_date = parse_excel_date(item.get('expiration_date'))
            if expiration_date is None and validity_months is not None:
                expiration_date = calc_expiration(
                    training_date,
                    validity_months=validity_months,
                    validity_days=None,
                )
            if validity_months is None and expiration_date is None:
                validity_months = 12
                expiration_date = calc_expiration(training_date, validity_months=12)

            trainings_to_create.append(Training(
                employee=employee,
                title=title,
                category=_cell_str(item.get('category')),
                training_date=training_date,
                trainer=_cell_str(item.get('trainer')),
                validity_months=validity_months,
                validity_days=None,
                expiration_date=expiration_date,
                process_classification=map_classification(item.get('process_classification')),
                remarks=_cell_str(item.get('remarks')),
                worker_line_status='Floating',
                take=take,
                created_by_id=getattr(user, 'id', None),
            ))

        created_count = 0
        if trainings_to_create:
            # bulk_create in batches; ignore_conflicts not needed due to de-dupe above
            for i in range(0, len(trainings_to_create), 1000):
                batch = trainings_to_create[i:i + 1000]
                Training.objects.bulk_create(batch, batch_size=1000)
                created_count += len(batch)

    log_audit(
        user,
        'IMPORT',
        'trainings',
        None,
        (
            f'Imported trainings: {created_count} created, '
            f'{len(to_create)} employees added, {len(to_update)} employees updated, '
            f'{skipped_duplicates} duplicates skipped, {len(errors)} row errors'
        ),
    )

    return {
        'rows_read': len(rows),
        'employees_created': len(to_create),
        'employees_updated': len(to_update),
        'trainings_created': created_count,
        'duplicates_skipped': skipped_duplicates,
        'errors': errors[:50],
        'error_count': len(errors),
    }
